import os
import openpyxl

# Step 1: Define the file path
file_path = r"C:\Users\gs235\Downloads\Barcodes.xlsx"

# Check if the file exists
if not os.path.exists(file_path):
    print(f"Error: The file '{file_path}' was not found. Please check the file path.")
else:
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # Assuming you want to work with the first sheet

    # Example processing: Check for "Vendor" in column A and delete rows
    rows_to_delete = [row for row in range(1, ws.max_row + 1) if ws[f'A{row}'].value == "Vendor"]

    print(f"Rows to delete: {rows_to_delete}")

    # Delete rows from bottom to top to prevent shifting
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)
        print(f"Deleted row {row}")

    # Step 2: Delete columns E (5th), F (6th), G (7th), H (8th)
    for i in range(4):
        ws.delete_cols(5)
        print(f"Deleted column {chr(69 + i)}")  # E is 69 in ASCII

    print("Workbook loaded.")

    # Dictionary to store item numbers and corresponding reference codes
    item_references = {}

    # Step 2: Iterate through each row and collect reference codes by item number
    for row in range(2, ws.max_row + 1):  # Assuming the first row is headers
        item_number = ws[f'C{row}'].value  # Get the item number from column C
        reference_code = ws[f'D{row}'].value  # Get the reference code from column D

        if item_number is not None:
            # If the item number exists in the dictionary, append the reference code
            if item_number in item_references:
                item_references[item_number].append(reference_code)
            else:
                # Otherwise, create a new entry for the item number
                item_references[item_number] = [reference_code]

    # Step 3: Clear the existing rows (starting from row 2) to avoid duplication
    ws.delete_rows(2, ws.max_row)

    # Step 4: Write back the consolidated data
    row_to_write = 2  # Start from the second row to overwrite data

    for item_number, references in item_references.items():
        # Write the item number back to column C
        ws[f'C{row_to_write}'] = item_number
        
        # Write each reference code into consecutive columns starting from column D
        for i, ref in enumerate(references):
            ws.cell(row=row_to_write, column=4 + i).value = ref  # Write reference codes starting from column D

        # Move to the next row after writing all reference codes for an item
        row_to_write += 1

    # Save the modified file
    wb.save(file_path)
    print(f"Item numbers and reference codes have been consolidated into separate cells. Changes saved to '{file_path}'.")
