from datetime import datetime
import os
import openpyxl

# Step 1: Define the file paths
file_path = r"C:\Users\gs235\Downloads\Final Inventory.xlsx"
new_file_path = r"C:\Users\gs235\Downloads\Sorted Final Inventory.xlsx"  # New sorted file

# Check if the original file exists
if not os.path.exists(file_path):
    print(f"Error: The file '{file_path}' was not found.")
else:
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # Assuming you want to work with the first sheet

    # Step 2: Extract all rows except the header
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the headers
        data.append(row)

    # Step 3: Sort the data by the expiration date in column E (index 4)
    data.sort(key=lambda row: datetime.strptime(str(row[4]), "%Y-%m-%d") if row[4] else datetime.max)

    # Step 4: Create a new workbook and add the sorted data
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # Step 5: Copy the header from the original file
    header = [cell.value for cell in ws[1]]  # Assuming the first row contains headers
    new_ws.append(header)

    # Step 6: Write the sorted data to the new worksheet
    for row_data in data:
        new_ws.append(row_data)

    # Step 7: Save the new workbook with the sorted data
    new_wb.save(new_file_path)

    print(f"Data has been successfully sorted by expiration date and saved to '{new_file_path}'.")
