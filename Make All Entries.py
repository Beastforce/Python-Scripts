from datetime import datetime
import os
import openpyxl

# Step 1: Define the file paths
file_path = r"C:\Users\gs235\Downloads\Barcodes.xlsx"
file_path_2 = r"C:\Users\gs235\Downloads\Export Cathay Data.xlsx"
file_path_3 = r"C:\Users\gs235\Downloads\Item Sheet.xlsx"
completed_file_path = r"C:\Users\gs235\Downloads\Completed.xlsx" 
completed_file_path_2 = r"C:\Users\gs235\Downloads\Completed_2.xlsx"  
completed_file_path_3 = r"C:\Users\gs235\Downloads\Completed_3.xlsx"  
combined_file_path = r"C:\Users\gs235\Downloads\Final Inventory.xlsx"

# Check if the files exist
if not os.path.exists(file_path) or not os.path.exists(file_path_2) or not os.path.exists(file_path_3):
    print(f"Error: One or more files were not found. Please check the file paths.")
else:
    # Load the workbooks
    wb = openpyxl.load_workbook(file_path)
    wb2 = openpyxl.load_workbook(file_path_2)
    wb3 = openpyxl.load_workbook(file_path_3)
    ws = wb.active  # First workbook's active sheet (Barcodes.xlsx)
    ws2 = wb2.active  # Second workbook's active sheet (Export Cathay Data.xlsx)
    ws3 = wb3.active  # Third workbook's active sheet (Item Sheet.xlsx)

    # Step 2: Create new workbooks for the completed data
    wb_completed = openpyxl.Workbook()
    ws_completed = wb_completed.active

    wb_completed_2 = openpyxl.Workbook()
    ws_completed_2 = wb_completed_2.active

    wb_completed_3 = openpyxl.Workbook()
    ws_completed_3 = wb_completed_3.active

    wb_combined = openpyxl.Workbook()
    ws_combined = wb_combined.active

    # Example processing: Check for "Vendor" in column A and delete rows
    rows_to_delete = [row for row in range(1, ws.max_row + 1) if ws[f'A{row}'].value == "Vendor"]

    print(f"Rows to delete: {rows_to_delete}")

    # Delete rows from bottom to top to prevent shifting
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)
        print(f"Deleted row {row}")

    # Step 3: Delete columns E (5th), F (6th), G (7th), H (8th)
    for i in range(4):
        ws.delete_cols(5)
        print(f"Deleted column {chr(69 + i)}")  # E is 69 in ASCII

    print("Workbook processing complete.")

    # Dictionary to store item numbers and corresponding reference codes
    item_references = {}

    # Step 4: Iterate through each row and collect reference codes by item number
    for row in range(2, ws.max_row + 1):  # Assuming the first row is headers
        item_number = ws[f'C{row}'].value  # Get the item number from column C
        reference_code = ws[f'D{row}'].value  # Get the reference code from column D

        if item_number is not None:
            # If the item number exists in the dictionary, append the reference code
            if item_number in item_references:
                item_references[item_number].append(reference_code)
            else:
                # Otherwise, create a new entry for the item number
                item_references[item_number] = [reference_code]

    # Step 5: Write the processed data to the new workbook
    row_to_write = 1  # Start from the first row in the new file

    # Create dynamic headers based on the maximum number of reference codes
    max_references = max(len(refs) for refs in item_references.values())
    headers = ["Item Number"] + [f"Reference Code {i+1}" for i in range(max_references)]

    # Write headers to the first row
    ws_completed.append(headers)

    # Write data below headers
    for item_number, references in item_references.items():
        # Create a row starting with the item number
        row_data = [item_number] + references
        
        # Append the row to the new worksheet
        ws_completed.append(row_data)

    # Write headers to the second completed file
    ws_completed_2.append(["Item Number", "Case Size", "Name", "Exp Date"])

    # Step 3: Iterate through each row and extract data from columns A, C, D, and the Exp Date column
    for row in ws2.iter_rows(min_row=2, values_only=True):  # Assuming the first row is headers
        # Extract the relevant columns
        item_number = row[0]  # Column A (Item Number)
        case_size = row[2]    # Column C (Case Size)
        name = row[3]         # Column D (Name)
        exp_date = row[14]    # Assuming the Exp Date is in column E (adjust index if needed)

        # Convert exp_date to string if it is a datetime object
        if isinstance(exp_date, datetime):
            exp_date = exp_date.strftime("%Y-%m-%d")  # Format the date as YYYY-MM-DD

        # Append the extracted data to the new worksheet
        ws_completed_2.append([item_number, case_size, name, exp_date])

    # Write headers to the completed worksheet if needed
    ws_completed_3.append(["Item Number", "Shelf Number"])

    # Step 1: Extract Data from ws3 and append to ws_completed_3
    for row in ws3.iter_rows(min_row=2, values_only=True):  # Assuming the first row is headers
        item_number = row[0]  # Column A (Item Number)
        shelf_number = row[1]  # Column B (Shelf Number)
        
        # Append the extracted data to ws_completed_3
        ws_completed_3.append([item_number, shelf_number])

    # Step 2: Remove "BULK-1" from column B
    for row in ws_completed_3.iter_rows(min_row=2, max_col=2):  # Start from row 2, limit to columns A and B
        if row[1].value == "BULK-1":  # Check if column B contains "BULK-1"
            row[1].value = None  # Clear the value in column B
            print(f"Cleared 'BULK-1' in row {row[0].row}")

    # Step 3: Track seen item numbers and remove rows with blank shelf locations
    seen_item_numbers = set()
    rows_to_delete = []

    # Iterate through each row in ws_completed_3
    for row in ws_completed_3.iter_rows(min_row=2, max_col=2):  # Limit to columns A and B
        item_number = row[0].value  # Column A (Item Number)
        shelf_location = row[1].value  # Column B (Shelf Location)

        # Check if the item number has been seen before
        if item_number in seen_item_numbers:
            # If item number has been seen and shelf location is blank, mark row for deletion
            if shelf_location is None:
                rows_to_delete.append(row[0].row)  # Collect row number for deletion
        else:
            # If it's the first occurrence of the item number, add it to the set
            seen_item_numbers.add(item_number)

    # Step 4: Delete marked rows from bottom to top
    for row in reversed(rows_to_delete):
        ws_completed_3.delete_rows(row)
    
    # Step 6: Save the completed data to new files
    wb_completed.save(completed_file_path)
    wb_completed_2.save(completed_file_path_2)
    wb_completed_3.save(completed_file_path_3)

    # Combine the completed files into one
    # Append data from Completed.xlsx

    # Step 1: Set up headers
    ws_combined.append(["Item Number", "Shelf Location", "Case Size", "Name", "Exp Date YYYY-MM-DD"])

    # Step 2: Append item numbers and shelf locations from ws_completed_3 into ws_combined
    for row in ws_completed_3.iter_rows(min_row=2, values_only=True):  # Skip headers
        item_number = row[0]  # Column A (Item Number)
        shelf_location = row[1]  # Column B (Shelf Location)
        
        # Append the extracted data to ws_combined
        ws_combined.append([item_number, shelf_location])

    # Step 3: Store data from ws_completed_2 (Case Size, Name, Exp Date) into a dictionary by item number
    data_case_info = {}  # Rename this dictionary to store case information

    for row in ws_completed_2.iter_rows(min_row=2, values_only=True):  # Skip headers
        item_number = row[0]  # Column A in ws_completed_2 (Item Number)
        case_size = row[1]    # Column C in ws_completed_2 (Case Size)
        name = row[2]         # Column D in ws_completed_2 (Name)
        exp_date = row[3]     # Assuming Exp Date is in column E

        # Convert exp_date to string if it's a datetime object
        if isinstance(exp_date, datetime):
            exp_date = exp_date.strftime("%Y-%m-%d")
        
        # Store the data in the dictionary by item number
        data_case_info[item_number] = [case_size, name, exp_date]

    # Step 4: Match item numbers in ws_combined with the data in ws_completed_2
    for row in ws_combined.iter_rows(min_row=2):  # Iterate over ws_combined, skipping headers
        item_number = row[0].value  # Get the item number from column A in ws_combined

        # Check if this item number exists in the data from ws_completed_2
        if item_number in data_case_info:
            case_size, name, exp_date = data_case_info[item_number]
            
            # Write the matched data to columns C, D, and E in ws_combined
            ws_combined.cell(row=row[0].row, column=3).value = case_size  # Column C (Case Size)
            ws_combined.cell(row=row[0].row, column=4).value = name       # Column D (Name)
            ws_combined.cell(row=row[0].row, column=5).value = exp_date   # Column E (Exp Date)

    start_col = 6  # Column 6 corresponds to column F

    # Step 5: Build a second dictionary from ws_completed using item numbers as keys for reference numbers
    data_reference_numbers = {}  # Dictionary to store reference numbers for each item number

    for row in ws_completed.iter_rows(min_row=2, values_only=True):  # Skip headers
        item_number = row[0]  # Column A in ws_completed (Item Number)
        references = list(row[1:])  # Collect all references in the row starting from column B onward
        
        # Store all references for each item number
        data_reference_numbers[item_number] = references

    # Step 6: Compare item numbers in ws_combined and append reference numbers to the matching rows
    for row in ws_combined.iter_rows(min_row=2):  # Iterate over ws_combined, skipping headers
        item_number = row[0].value  # Get the item number from column A in ws_combined

        # Check if this item number exists in the data from ws_completed
        if item_number in data_reference_numbers:
            reference_numbers = data_reference_numbers[item_number]
            
            # Start writing the reference numbers from column F (index 6) onward in ws_combined
            for i, reference_number in enumerate(reference_numbers):
                ws_combined.cell(row=row[0].row, column=start_col + i).value = reference_number

    # Step 7: Save the combined data into the new file
    wb_combined.save(combined_file_path)

    print(f"Data from the completed files has been combined and saved to '{combined_file_path}'.")