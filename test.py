import os
import openpyxl

# Step 1: Define the file path
file_path = r"C:\Users\gs235\Downloads\Barcodes.xlsx"

# Check if the file exists
if not os.path.exists(file_path):
    print(f"Error: The file '{file_path}' was not found. Please check the file path.")
else:
    # Proceed with loading and processing the workbook if the file exists
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # Assuming you want to work with the first sheet

    # Example processing: Check for "Vendor" in column A and delete rows
    rows_to_delete = [row for row in range(1, ws.max_row + 1) if ws[f'A{row}'].value == "Vendor"]

    # Delete rows from bottom to top to prevent shifting
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)

    # Step 2: Delete columns E (5th), F (6th), G (7th), H (8th)
    ws.delete_cols(5)  # Deletes column E
    ws.delete_cols(5)  # After E is deleted, F becomes the new 5th column, so delete it
    ws.delete_cols(5)  # G becomes the new 5th column, so delete it
    ws.delete_cols(5)  # H becomes the new 5th column, so delete it

    # Save the modified file
    wb.save(file_path)
    print(f"Rows containing 'Vendor' in column A have been deleted. Changes saved to '{file_path}'.")