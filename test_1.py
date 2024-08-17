import os
import openpyxl

# Step 1: Define the file path
file_path = r"C:\Users\gs235\Downloads\Barcodes.xlsx"

print("Starting script...")

# Check if the file exists
if not os.path.exists(file_path):
    print(f"Error: The file '{file_path}' was not found. Please check the file path.")
else:
    print(f"File '{file_path}' found. Proceeding...")

    # Proceed with loading and processing the workbook if the file exists
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # Assuming you want to work with the first sheet
    print("Workbook loaded.")

    # Example processing: Check for "Vendor" in column A and delete rows
    rows_to_delete = [row for row in range(1, ws.max_row + 1) if ws[f'A{row}'].value == "Vendor"]

    print(f"Rows to delete: {rows_to_delete}")

    # Delete rows from bottom to top to prevent shifting
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)
        print(f"Deleted row {row}")

    # Step 2: Delete columns E (5th), F (6th), G (7th), H (8th)
    for i in range(4):
        ws.delete_cols(5)
        print(f"Deleted column {chr(69 + i)}")  # E is 69 in ASCII

    # Save the modified file
    wb.save(file_path)
    print(f"Rows containing 'Vendor' in column A have been deleted. Changes saved to '{file_path}'.")
